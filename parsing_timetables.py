import openpyxl
import json

def parser_faculty_tt():
    #json config file for parser
    with open('parser.json') as f:
      parserConfig = json.load(f)

    facultyTTWB = openpyxl.load_workbook(parserConfig["filenamefacultytt"])
    facultyTTSheet = facultyTTWB.active

    class_loc = list()
    for fac in range(0,int(parserConfig["facultycount"] ) ):
        initial = int( parserConfig["rowsperfacultytt"] ) * fac + 3
        for row in range(initial,initial+9 +1,2):
            for col in list(parserConfig["classcolumns"]):
                className = facultyTTSheet.cell(row=row,column=col)
                classLoc = facultyTTSheet.cell(row=row+1,column=col)
                class_loc.append( (className.value,classLoc.value) )
                #print(className.value,"<>",classLoc.value,"@",row)

    return class_loc, len( class_loc )

def parser_class_tt():
    #json config file for parser
    with open('parser.json') as f:
      parserConfig = json.load(f)

    classTTWB = openpyxl.load_workbook(parserConfig["filenameclasstt"])
    classTTSheet = classTTWB.active

    class_ = list()
    initial = 3
    friday = 11
    for row in range(initial,initial+9 +1,2):
        for col in list(parserConfig["classcolumns"]):
            if( row != friday):
                className = classTTSheet.cell(row=row,column=col)
            else:
                className = classTTSheet.cell(row=row+1, column=col)

            class_.append( className.value )
            #print(className.value,"@",row)

    return class_, len( class_ )

