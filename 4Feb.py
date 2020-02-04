from openpyxl import load_workbook

def isProjectPeriod(hr,day):
    projectperiodnamegiveninclasstt = ["PROJECT"]
    ret = False
    if( classtt.cell(column=classH[hr], row=workD[day]).value in projectperiodnamegiveninclasstt ):
        ret =  True
    if( hr > 2):
            ret = classtt.cell(column=classH[hr - 2], row=workD[day]).value in projectperiodnamegiveninclasstt or ret
    if (hr > 1):
         ret = classtt.cell(column = classH[hr - 1], row=workD[day]).value in projectperiodnamegiveninclasstt or ret
    return ret
def isGuideInBetaProjectPeriod(hr,day,guideslno):
    projectperiodnamegiveninguidestt = ["PROJECT"]
    className = ["S8 CS B","S8 CSB"]
    ret = False
    if (hr > 1):
        cellvalue1awayunder = guidestt.cell(column = classH[hr - 1], row=guideday(workD[day],guideslno)+1).value
        cellvalue1away = guidestt.cell(column = classH[hr - 1], row=guideday(workD[day],guideslno)).value
    if (hr > 2):
        cellvalue2away = guidestt.cell(column=classH[hr - 2], row=guideday(workD[day],guideslno)).value
        cellvalue2awayunder = guidestt.cell(column=classH[hr - 2], row=guideday(workD[day],guideslno)+1).value
    cellvalue = guidestt.cell(column=classH[hr], row=guideday(workD[day],guideslno)).value
    cellvalueunder = guidestt.cell(column=classH[hr], row=guideday(workD[day],guideslno)+1).value
    if((cellvalue in projectperiodnamegiveninguidestt) and ( cellvalueunder in className )):
        ret = True
  #      print("ok",ret)
    if( hr > 2):
            ret = ((( cellvalue2away in projectperiodnamegiveninguidestt) and ( cellvalue2awayunder in className )) or ret)
   #         print("Hereeee", ret)
    if (hr > 1):
         ret = ((( cellvalue1away in projectperiodnamegiveninguidestt) and ( cellvalue1awayunder in className)) or ret)
    #     print("Herer",ret)
    return ret

def twoCellsWithinPeriod(hr,day,guideslno):
    period = set()
    cellvalue1away = "XX"
    cellvalue2away = "XX"
    if (hr > 1):
        cellvalue1away = guidestt.cell(column=classH[hr - 1], row=guideday(workD[day], guideslno)).value
    if (hr > 2):
        cellvalue2away = guidestt.cell(column=classH[hr - 2], row=guideday(workD[day], guideslno)).value
    cellvalue = guidestt.cell(column=classH[hr], row=guideday(workD[day], guideslno)).value
    period = {cellvalue1away,cellvalue2away,cellvalue}
    return period

row_workD = { 3:1,5:2,7:3,9:4,12:5}
column_classH = {3:1,4:2,6:3,7:4,9:5,11:6,12:7}

def isFreePeriod(hr,day,guideslno=1):
    # print("#",guidestt.cell(column=hr, row=guideday(day,guideslno) ).value)
    around = twoCellsWithinPeriod(column_classH[hr],row_workD[day],guideslno)
    return ( guidestt.cell(column=hr, row=guideday(day,guideslno) ).value is  None ) and ( around.intersection({"PROJECT"}) == {} )
def guideday(day,slno=1):
    return (slno-1)*15 + day

FILENAME1 = "_class_tt_org.xlsx"
SHEETNAME1 = "Sheet1"
FILENAME2 = "_guide_tt_org.xlsx"
SHEETNAME2 = "Sheet1"

workbook1 = load_workbook(filename=FILENAME1)
workbook2 = load_workbook(filename=FILENAME2)


classtt = workbook1[SHEETNAME1]
guidestt = workbook2[SHEETNAME2]

#day = 1
#hr = 6
workD = { 1:3,2:5,3:7,4:9,5:12}
classH = {1:3,2:4,3:6,4:7,5:9,6:11,7:12}

week = 0
startDay = 1
endDay = 5
startHour = 1
endHour = 7
"""
teams as given in S8 details file, Meenu Mathew Group Ignored while numbering and assigning 
total allocated groups 18 in number
"""
team_guide_slno = {1:25,2:19,3:5,4:13,5:8,6:19,7:12,8:16,9:9,10:9,11:6,12:16,13:4,14:4,15:23,16:17,17:29,18:29}
allocated = set()
notAllocated = [1,2,3,4,5,6,7,8,9,10,11,12,13,14,15,16,17,18]
while( allocated != {1,2,3,4,5,6,7,8,9,10,11,12,13,14,15,16,17,18} and week <= 1 ):
    week = week + 1
    for day in range(startDay,endDay+1):
        for hr in range(startHour,endHour+1):
            if( isProjectPeriod(hr,day) ):
                before = allocated
                for team in notAllocated:
                    if( isGuideInBetaProjectPeriod(hr,day,team_guide_slno[team]) ):
                        """
                        allocating team with its guide having project hour in cse B
                        """
                        allocated.add(team)
                        notAllocated.remove(team)
                        print("team:?",team,", allocated on:",hr,day,week)
                        break
                    elif( isFreePeriod(classH[hr],workD[day],team_guide_slno[team])):
                        """
                        allocating as teams guide is free
                        """
                        allocated.add(team)
                        notAllocated.remove(team)
                        print("team:!", team, ", allocated on:", hr, day, week)
                        break
                    else:
                        pass
                after = allocated
                if(after == before):
                    print("Allocation for no team possible in hour,day,week",hr,day,week)
