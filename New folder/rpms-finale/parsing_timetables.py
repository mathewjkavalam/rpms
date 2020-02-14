import openpyxl
import json
from os import path

def parser_faculty_tt():
    #json config file for parser
    with open('parser.json') as f:
      parserConfig = json.load(f)

    facultyTTWB = openpyxl.load_workbook(parserConfig["filenamefacultytt"])
    facultyTTSheet = facultyTTWB.active

    class_loc = list()
    for fac in range(0,int(parserConfig["facultycount"] ) ):
        initial = int( parserConfig["rowsperfacultytt"] ) * fac + 3
        for row in range(initial,initial+9 +1,2):
            for col in list(parserConfig["classcolumns"]):
                className = facultyTTSheet.cell(row=row,column=col)
                classLoc = facultyTTSheet.cell(row=row+1,column=col)
                class_loc.append( (className.value,classLoc.value) )
                # print(className.value,"<>",classLoc.value,"@",row)

    return class_loc, len( class_loc )

def parser_class_tt():
    #json config file for parser
    with open('parser.json') as f:
      parserConfig = json.load(f)

    classTTWB = openpyxl.load_workbook(parserConfig["filenameclasstt"])
    classTTSheet = classTTWB.active

    class_ = list()
    initial = 3
    friday = 11
    for row in range(initial,initial+9 +1,2):
        for col in list(parserConfig["classcolumns"]):
            if( row != friday):
                className = classTTSheet.cell(row=row,column=col)
            else:
                className = classTTSheet.cell(row=row+1, column=col)

            class_.append( className.value )
            #print(className.value,"@",row)

    return class_, len( class_ )
#print( parser_faculty_tt() )

def merged_cell_handler_faculty_tt():
    with open('parser.json') as f:
        parserConfig = json.load(f)

    threelongperiod = ["PROJECT","FOSS LAB"]
    twolongperiod = ["X"]
    #project periods,lab periods
    tt = parser_faculty_tt()[0]
    with open('parser.json') as f:
      parserConfig = json.load(f)

    for fac in range( int(parserConfig["facultycount"]) ):
        for day in range(5):
            exclude = 0
            for hr in range(7):
                idx = fac*35 + day*7 + hr
                if( exclude > 0 ):
                    exclude = exclude - 1
                elif( tt[idx][0] in threelongperiod and exclude <= 0):
                    tt[idx+1] = tt[idx]
                    tt[idx + 2] = tt[idx]
                    exclude = 2
                elif( tt[idx][0] in twolongperiod and exclude <= 0):
                    tt[idx + 1] = tt[idx]
                    exclude = 1
                elif (tt[idx][0] in [None]):
                    tt[idx] = ("FREE", "")


    out = parserConfig["id"]+"merged_cell_handler_faculty_tt"+".json"
    if(not path.exists(out) ):
        with open( out , 'w') as outfile:
            outfile.write(json.dumps(tt))
    return path.exists(out)

def merged_cell_handler_class_tt():
    with open('parser.json') as f:
        parserConfig = json.load(f)

    threelongperiod = ["PROJECT","FOSS LAB"]
    twolongperiod = ["X"]
    tt = parser_class_tt()[0]

    for day in range(5):
        exclude = 0
        for hr in range(7):
            idx = day*7 + hr
            if( exclude > 0 ):
                exclude = exclude - 1
            elif( tt[idx] in threelongperiod and exclude <= 0):
                tt[idx+1] = tt[idx]
                tt[idx + 2] = tt[idx]
                exclude = 2
            elif( tt[idx] in twolongperiod and exclude <= 0):
                tt[idx + 1] = tt[idx]
                exclude = 1
            elif (tt[idx][0] in [None]):
                tt[idx] = "FREE"

    out = parserConfig["id"] + "merged_cell_handler_class_tt" + ".json"
    if ( not path.exists(out)):
        with open(out, 'w') as outfile:
            outfile.write(json.dumps(tt))
    return path.exists(out)
"""
Pwoli Sharath!!!
"""