from openpyxl import load_workbook

"""
Input:
Ecxel Files
...
Input:
Multiplicity = Number of Teams per Project Period
Start Date
Panel Members
"""
def isGuideIn3(hr,day,guideslno,className = ["S8 CS B","S8 CSB"],periodnamegiveninguidestt = ["PROJECT"]):
    ret = False
    if (hr > 1):
        cellvalue1awayunder = guidestt.cell(column = classH[hr - 1], row=guideday(workD[day],guideslno)+1).value
        cellvalue1away = guidestt.cell(column = classH[hr - 1], row=guideday(workD[day],guideslno)).value
    if (hr > 2):
        cellvalue2away = guidestt.cell(column=classH[hr - 2], row=guideday(workD[day],guideslno)).value
        cellvalue2awayunder = guidestt.cell(column=classH[hr - 2], row=guideday(workD[day],guideslno)+1).value
    cellvalue = guidestt.cell(column=classH[hr], row=guideday(workD[day],guideslno)).value
    cellvalueunder = guidestt.cell(column=classH[hr], row=guideday(workD[day],guideslno)+1).value
    if((cellvalue in periodnamegiveninguidestt) and ( cellvalueunder in className )):
        ret = True
    if( hr > 2):
            ret = ((( cellvalue2away in periodnamegiveninguidestt) and ( cellvalue2awayunder in className )) or ret)
    if (hr > 1):
         ret = ((( cellvalue1away in periodnamegiveninguidestt) and ( cellvalue1awayunder in className)) or ret)
    return ret

def guideday(day,slno=1):
    return (slno-1)*15 + day

def isFreePeriod(hr,day,guideslno=1):
    around3 = twoCellsWithinPeriod(hr,day,guideslno)
    around2 = oneCellWithinPeriod(hr,day,guideslno)
    return ( guidestt.cell(column=classH[hr], row=guideday( workD[day],guideslno) ).value is  None ) and ( around3.intersection(long3Periods) == set() ) and ( around2.intersection(long2Periods) == set() )

def twoCellsWithinPeriod(hr,day,guideslno):
    period = set()
    cellvalue1away = "XX"
    cellvalue2away = "XX"
    if (hr > 1):
        cellvalue1away = guidestt.cell(column=classH[hr - 1], row=guideday(workD[day], guideslno)).value
    if (hr > 2):
        cellvalue2away = guidestt.cell(column=classH[hr - 2], row=guideday(workD[day], guideslno)).value
    cellvalue = guidestt.cell(column=classH[hr], row=guideday(workD[day], guideslno)).value
    period = {cellvalue1away,cellvalue2away,cellvalue}
    return period
def oneCellWithinPeriod(hr,day,guideslno):
    period = set()
    cellvalue1away = "XX"
    if (hr > 1):
        cellvalue1away = guidestt.cell(column=classH[hr - 1], row=guideday(workD[day], guideslno)).value
    cellvalue = guidestt.cell(column=classH[hr], row=guideday(workD[day], guideslno)).value
    period = {cellvalue1away,cellvalue}
    return period


FILENAME2 = "_guide_tt_org.xlsx"
SHEETNAME2 = "Sheet1"
workbook2 = load_workbook(filename=FILENAME2)
guidestt = workbook2[SHEETNAME2]


workD = { 1:3,2:5,3:7,4:9,5:12}
classH = {1:3,2:4,3:6,4:7,5:9,6:11,7:12}
column_classH = {3:1,4:2,6:3,7:4,9:5,11:6,12:7}
long3Periods = set()
for p in ["PROJECT","MICROPROCESSOR LAB","C LAB","CP LAB","DIGITAL LAB","MP LAB","NW LAB","FOSS LAB","N/W LAB","MINI PROJECT"]:
    long3Periods.add(p)
long2Periods = set()
for p in ["COMPREHENSIVE"]:
    long2Periods.add(p)

faculty = list()

for g in range(1,18+1):
    ft = []
    faculty.append(ft)
    for d in [1,2,3,4,5]:
        for h in [1,2,3,4,5,6,7]:
            #print(h,d,g)
            if( isGuideIn3(h, d, g,["S8 CS A","S8 CSA"]) ):
                ft.append( (h,d,'A') )

            elif( isGuideIn3(h, d, g,["S8 CS B","S8 CSB"]) ):
                ft.append( (h,d,'B') )

            elif( isGuideIn3(h, d, g,["S8 CS C","S8 CSC"]) ):
                ft.append( (h,d,'C') )
            elif( isFreePeriod(h,d,g) ):
                ft.append((h, d, 'FREE'))
    faculty[g - 1] = tuple(ft)
print(faculty[1-1] )