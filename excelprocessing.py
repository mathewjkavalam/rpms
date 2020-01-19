from openpyxl import load_workbook
FILENAME = "timetableguide.xlsx"
SHEETNAME = "Sheet1"
workbook = load_workbook(filename=FILENAME)
#print("#",workbook.sheetnames)
worksheet = workbook[SHEETNAME]
"""
merged row cells need to take the starting row
to retrieve value;otherwise its none.
output:
# S1(D3-G3)
# None
# None
# None
"""
#COLUMN = 1
#ROW = 3
#print("#",worksheet.cell(column=COLUMN,row=ROW).value)
name = (1,1)

mon = 3
tue = 5
wed = 7
thu = 9
fri = 11

COLUMN = 1
ROW = 0

classhours = [0,3,4,6,7,9,11,12]
#for COLUMN in range(4,7+1):
#    print("#",worksheet.cell(column=COLUMN,row=ROW).value)
"""
Sample INPUT:
(fri,5)
Sample OUTPUT:
S7
"""
print("#",worksheet.cell(column=classhours[5],row=fri).value)