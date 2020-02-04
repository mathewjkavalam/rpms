"""
before integration 03/Feb/2020
"""
from openpyxl import load_workbook

def isProjectPeriod(hr,day):
    projectperiodnamegiveninclasstt = ["PROJECT"]
    ret = False
    if( classtt.cell(column=classH[hr], row=workD[day]).value in projectperiodnamegiveninclasstt ):
        ret =  True
    if( hr > 2):
            ret = classtt.cell(column=classH[hr - 2], row=workD[day]).value in projectperiodnamegiveninclasstt or ret
    if (hr > 1):
         ret = classtt.cell(column = classH[hr - 1], row=workD[day]).value in projectperiodnamegiveninclasstt or ret
    return ret
def isFreePeriod(hr,day,guideslno=1):
    # print("#",guidestt.cell(column=hr, row=guideday(day,guideslno) ).value)
    return guidestt.cell(column=hr, row=guideday(day,guideslno) ).value is  None
def guideday(day,slno=1):
    return (slno-1)*15 + day


"""
Get Name of Faculty from _guide_tt.xlxs
print( worksheet2.cell(column=1, row= guideday(3, slno) ).value )

"""

"""
INPUT:
    
    FILES:
    class timetable
    faculty time table
    
    VARIABLE:
    day
    hr
    number of times each faculty is called 
    project coordinator name .
    team guide name
    number of total panel members
OUTPUT:
    panel Members Name
"""
FILENAME1 = "_class_tt_org.xlsx"
SHEETNAME1 = "Sheet1"
FILENAME2 = "_guide_tt_org.xlsx"
SHEETNAME2 = "Sheet1"

workbook1 = load_workbook(filename=FILENAME1)
workbook2 = load_workbook(filename=FILENAME2)


classtt = workbook1[SHEETNAME1]
guidestt = workbook2[SHEETNAME2]

# mon six hour
day = 1
hr = 6

workD = { 1:3,2:5,3:7,4:9,5:12}
classH = {1:3,2:4,3:6,4:7,5:9,6:11,7:12}
#initialising
CountOfFacultyCalled = {}
for i in range(1,32+1):
        CountOfFacultyCalled[i] = 0
#CountOfFacultyCalled = {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 6: 0, 7: 0, 8: 0, 9: 0, 10: 0, 11: 0, 12: 0, 13: 0, 14: 0, 15: 0, 16: 0, 17: 0, 18: 0, 19: 0, 20: 0, 21: 0, 22: 0, 23: 0, 24: 0, 25: 0, 26: 0, 27: 0, 28: 0, 29: 0, 30: 0, 31: 0, 32: 0}
coordname = "Amitha Mathew"
guidename = "Hari Sir"
maxpanelcount = 2
# print("7879",isProjectPeriod(1,1))
"""
check if coord free 
"""
def coordnameSlno(name = "" ):
    return 19
print( "Coordinator Available:",isFreePeriod( classH[hr],workD[day],coordnameSlno(name=coordname) ) )

"""
check if guide free :)
"""
def guideSlno(name = "" ):
    return 25
if( guidename != coordname ):
    print( "Guide Available", isFreePeriod( classH[hr],workD[day],guideSlno(name=guidename) ) )
else:
    print("Guide and Coordinator same")
"""
Order remaining faculties
Priority Criteria:
    1 should have free on that hour 
    2 arrange ascendingly count called
    3 take the one which is available only now(overiding even second signal)  
"""
MAX_FACULTY_COUNT = 32

faculty = set()
for slno in range(1,MAX_FACULTY_COUNT+1):
    faculty.add(slno)
remaingfaculty = faculty-{guideSlno(guidename),coordnameSlno(coordname)}

# print("RemaingFacultyInitial",remaingfaculty)

# print(remaingfaculty)
# print(faculty)
# facultyScore[1] = {"free":False,"calledcount":8,"freecount":3}
# print(facultyScore[1]["freecount"])
facultyScore = {}
for num in remaingfaculty:
    if( isFreePeriod(classH[hr],workD[day],num) is False ):
        remaingfaculty = remaingfaculty - {num}
print("RemaingFacultyInitial",remaingfaculty)

# print(remaingfaculty)
# print(facultyScore)
"""
sorting 
"""

calledcount = {}
if( len(remaingfaculty) != 0 ):
    for num in remaingfaculty:
        facultyScore[num] = {"calledcount":CountOfFacultyCalled[num] , "freecount":0}
        try:
            calledcount[CountOfFacultyCalled[num]].add(num)
        except(KeyError):
            calledcount[CountOfFacultyCalled[num]] = {num}
print(facultyScore)
print(calledcount)
"""
sorting end
"""
overriders = 0
for fac in remaingfaculty:
    for day in range(1, 5+1):
        for hr in range(1, 7+1):
            if( isProjectPeriod(hr,day) and isFreePeriod(classH[hr],workD[day],fac) ):
                facultyScore[fac]["freecount"] = facultyScore[fac]["freecount"] + 1
                # print("free",fac,facultyScore[fac]["freecount"])
    if facultyScore[fac]["freecount"] == 1:
        # print("Overriders")
        overriders = overriders + 1
print(facultyScore)
""""
overriding
"""
allocated = {guideSlno(guidename),coordnameSlno(coordname)}
while overriders > 0 and len(allocated) < maxpanelcount and len(remaingfaculty) != 0:
    for fac in remaingfaculty:
        if facultyScore[fac]["freecount"] == 1 and len(allocated) < maxpanelcount:
            #Allocating
            --overriders
            print("Overriding")
            remaingfaculty = remaingfaculty - {fac}
            allocated.add(fac)
            #updating
            CountOfFacultyCalled[fac] = CountOfFacultyCalled[fac] + 1

print(calledcount)
while(len(allocated) < maxpanelcount and len(remaingfaculty) != 0):
    print("Allo",allocated)
    print("Remfa",remaingfaculty)
    for num in sorted(calledcount):
        print("Num",num)
        for n in calledcount[num]:
            if(len(allocated) < maxpanelcount):
                remaingfaculty = remaingfaculty - {n}
                allocated.add(n)
                CountOfFacultyCalled[n] = CountOfFacultyCalled[n] + 1
                calledcount[num] = calledcount[num] - {n}
if( len(remaingfaculty) != 0 and len(allocated) < maxpanelcount):
    print("UNsucess")
if( len(remaingfaculty) == 0 and len(allocated) < maxpanelcount):
    print("Not Possible")
if( len(remaingfaculty) >= 0 and len(allocated) >= maxpanelcount):
    print("Possible")
print("Allocated",allocated)
print("CountOfFacultyCalled(Copy to Next iteration)",CountOfFacultyCalled)