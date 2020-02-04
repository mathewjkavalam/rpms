from openpyxl import load_workbook
"""
from DB:
    excelsheets(2):
       <1> "_class_tt.xlsx"
            "Sheet1"
       <2> "_guide_tt.xlsx"
           "Sheet1"
    
"""

FILENAME1 = "_class_tt.xlsx"
SHEETNAME1 = "Sheet1"
FILENAME2 = "_guide_tt.xlsx"
SHEETNAME2 = "Sheet1"
workbook1 = load_workbook(filename=FILENAME1)
workbook2 = load_workbook(filename=FILENAME2)
classtt = workbook1[SHEETNAME1]
guidestt = workbook2[SHEETNAME2]
"""
standard:
    class timetable just like _class_tt 
"""
workD = { 1:3,2:5,3:7,4:9,5:12}
classH = {1:3,2:4,3:6,4:7,5:9,6:11,7:12}

"""
number of times faculty called read from DB.(retr)
"""
CountOfFacultyCalled = {1:0,2:0,3:1,4:0,5:0,6:0,7:0,8:0,9:0,10:0,11:0}


